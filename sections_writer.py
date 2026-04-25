from typing import AnyStr, Generator, List
import openpyxl
import polars as pl
import pandas as pd
import narwhals as nw
import ipaddress
import regex as re
from narwhals.typing import IntoDataFrameT
from openpyxl import load_workbook


class SectionsWriter:
    def __init__(self, dest_workbook_path: str) -> None:
        self.dest_workbook_path = dest_workbook_path
    
    def writer(self, dict_tionary: dict) -> None:
        pass
    
    

class Prefix_section:
    def __init__(self) -> None:        
        self.section_column_mapping = {
                "prefix_set_name*":	"Prefix-list Name",
                "ip_subnet": "Prefix-IP"
        }
        
        self.section_required_columns = [
            "Action",
            "Sequence Action",
            "Version",
            "Prefix-list Name",
            "Sequence Number",
            "Permit/Deny",
            "Prefix-IP",
            "Length",
            "Route Policy Mapping",
            "Access-list Protocol",
            "Access-list Source IP",
            "Access-list source Wild Mask",
            "Access-list Destination IP",
            "Access-list destination Wild Mask"
        ]
        
    
    @nw.narwhalify
    def length_column_handler_xr(self, df: IntoDataFrameT) -> List[AnyStr]:
        # 1. Define conditions for cleaner code
        has_eq = ~nw.col('expression_eq').is_null()
        has_ge = ~nw.col('expression_ge').is_null()
        has_le = ~nw.col('expression_le').is_null()

        # 2. Pre-cast to strings for safe concatenation
        str_eq = nw.col('expression_eq').cast(nw.String)
        str_ge = nw.col('expression_ge').cast(nw.String)
        str_le = nw.col('expression_le').cast(nw.String)

        # 3. Build the properly nested logic block
        formatted_expr = (
            nw.when(has_eq & (has_ge | has_le))
            .then(nw.lit("ERROR: Conflict between eq and ge/le"))
            .otherwise(
                nw.when(has_eq)
                .then(nw.lit("eq ") + str_eq)
                .otherwise(
                    nw.when(has_ge & has_le)
                    .then(nw.lit("ge ") + str_ge + nw.lit(" le ") + str_le)
                    .otherwise(
                        nw.when(has_ge)
                        .then(nw.lit("ge ") + str_ge)
                        .otherwise(
                            nw.when(has_le)
                            .then(nw.lit("le ") + str_le)
                            .otherwise(nw.lit(""))
                        )
                    )
                )
            )
        ).alias("Length")
        
        __processed_df = df.with_columns(formatted_expr)
        
        return __processed_df
    
    
    def get_ip_version(self, ip_value: str) -> str:
        if not isinstance(ip_value, str) or not ip_value.strip():
            return ""
        
        try:
            clean_ip = ip_value.strip().split('/')[0]
                        
            version = ipaddress.ip_address(clean_ip).version
            return f"ipv{version}"
        
        except ValueError:
            return ""
    
    
    @nw.narwhalify
    def get_the_version_xr(self, df: IntoDataFrameT) -> IntoDataFrameT:
        col_name = 'ip_subnet'
        
        result_df = nw.to_native(df)
        
        if isinstance(result_df, pd.DataFrame):
            result_df['Version'] = result_df[col_name].fillna('').astype(str).str.strip().map(self.get_ip_version)
            
        if isinstance(result_df, pl.DataFrame):
            result_df = result_df.with_columns(
                pl.col(col_name).fill_null("").cast(pl.String).map_elements(self.get_ip_version, return_dtype=pl.String).alias("Version")
            )
        result_df = nw.from_native(result_df)
        return result_df
        
    
    @nw.narwhalify
    # def action_sequence_action_columns_handler_xr(self, df: IntoDataFrameT) -> Tuple[List[AnyStr], List[AnyStr]]:
    def action_sequence_action_columns_handler_xr(self, df: IntoDataFrameT) -> IntoDataFrameT:
        op_col = nw.col("Operation").cast(nw.String)
        
        is_target_op = op_col.str.to_lowercase().str.contains('add|delete')
        
        expr_a = nw.lit("A:") + op_col.str.to_titlecase()
        
        expr_b = nw.when(
            is_target_op
        ).then(
            nw.lit(
                "S:"
            ) + op_col.str.to_titlecase()
        ).otherwise(
            nw.lit("")
        )
        
        result_df = df.with_columns(
            expr_a.alias("Action"),
            expr_b.alias("Sequence Action")
        )
        
        return result_df
    
    
    def section_writer(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, df: IntoDataFrameT, vendor_type: AnyStr="xr"):
        start_row = 1
        
        # worksheet = workbook_object["prefix"]
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        # worksheet_df = None
        
        if max_row != 1 and max_col != 1:
            worksheet.delete_rows(1, max_row)
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if vendor_type == 'xr':
            neo_df = self.action_sequence_action_columns_handler_xr(df)
            neo_df = self.length_column_handler_xr(neo_df)
            neo_df = self.get_the_version_xr(neo_df)
            
            keys_list = list(self.section_column_mapping.keys())
            
            for key in keys_list:
                value = self.section_column_mapping[key]
                neo_df= neo_df.with_columns(
                    neo_df[key].alias(value),
                )
                
            columns = [column for column in list(neo_df.columns) if column in self.section_required_columns]
            neo_df = neo_df[columns]
            
            col_dict = {
                column: indx for indx, column in enumerate(self.section_required_columns, 1)
            }
            
            section_required_column_mapping_with_df = [col_dict[column] for column in columns]
            
            for column_name, col_idx in col_dict.items():
                worksheet.cell(row=start_row, column=col_idx, value=column_name)
            
            start_row += 1
            
            # print(neo_df)
            for row_id in range(start_row, start_row + len(neo_df)):
                idx = row_id - start_row
                
                if str(neo_df[idx, 'Length']) == "ERROR: Conflict between eq and ge/le":
                    for col_id in range(1, len(self.section_required_columns) + 1):
                        if col_id < 3:
                            col_name = self.section_required_columns[col_id - 1]
                            worksheet.cell(row=row_id, column=col_id, value=neo_df[idx, col_name])
                        
                        else:
                            worksheet.cell(row=row_id, column=col_id, value="")
                            
                else:
                    for col_id in range(1, len(self.section_required_columns) + 1):
                        if col_id in section_required_column_mapping_with_df:
                            col_name = self.section_required_columns[col_id - 1]
                            worksheet.cell(row=row_id, column=col_id, value=neo_df[idx, col_name])
                            
                        
                        else:
                            worksheet.cell(row=row_id, column=col_id, value="")
                        
                        

class Policy_section:
    
    # ---- Regex patterns (compiled once at class load) ------------------------

    # Matches the control keyword at the start of a line: if / elseif / else
    _RE_CONTROL = re.compile(
        r"^\s*(elseif|else|if)\b", re.IGNORECASE
    )

    # End-of-block markers we must respect
    _RE_ENDIF = re.compile(r"^\s*endif\b", re.IGNORECASE)

    # "destination in <GROUP>" — tolerates optional parentheses & spacing
    _RE_DEST_IN = re.compile(
        r"destination\s+in\s+([\w\-.:]+)",
        re.IGNORECASE,
    )
    
    # Any other condition expression between the control keyword and "then"
    #   e.g.  if (community matches-any XYZ) then
    #         elseif extcommunity rt matches-any FOO then
    _RE_CONDITION_BODY = re.compile(
        r"^\s*(?:if|elseif)\s*\(?\s*(.*?)\s*\)?\s*then\s*$",
        re.IGNORECASE,
    )

    # Control keywords that terminate an action-accumulation phase
    _TERMINATORS = ("if", "elseif", "else", "endif")

    # Excel header row
    _HEADERS = (
        "Action",
        "Sequence/Set Action",
        "Version",
        "Route-Policy Name",
        "Sequence Number",
        "Permit/Deny",
        "Condition",
        "Match/If 'Destination in' Statement",
        "Match/If other than 'Destination in' Statement",
        "Set/Action Statement",
    )

    def __init__(self) -> None:
        self.section_column_mapping = {
                "name":	"Route-Policy Name",
        }
        
    
    @nw.narwhalify
    def pattern_matcher_extracter_and_writer_xr(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, df: IntoDataFrameT):
        # regex_pattern = r"(?:if|elseif)\s+\(?\s*(?:destination\s+in\s+)?(.*?)\s*\)?\s*then"
        # ending_string = "endif"
        # conditionals = ["if", "elseif", "else"]
        # else_regex_pattern = r"\A\belse\b\Z"
        
        # ---- Iterate policies ------------------------------------------------
        for row in df.iter_rows(named=True):
            policy_name = row.get("name") or ""
            lines = row.get("condition_split_list") or []

            if not lines:
                continue

            for parsed in self._parse_policy_blocks(lines):
                worksheet.append([
                    "",
                    "",
                    "",
                    policy_name,
                    "",
                    "",         
                    parsed["condition"],
                    parsed["dest_in"],
                    parsed["other_match"],
                    parsed["actions"],
                ])
                
    
    def _parse_policy_blocks(self, lines) -> Generator[dict, None, None]:
        """
        Walk a single policy's condition_split_list once and yield a dict
        per condition block:
            {condition, dest_in, other_match, actions}
        """
        current = None          # active conditional block being filled
        saw_control = False     # did we ever encounter if/elseif/else?
        standalone_actions = [] # collected action lines when no control seen

        for raw in lines:
            if raw is None:
                continue
            line = raw.strip()
            if not line:
                continue

            lower = line.lower()

            # ---- endif: flush and stop current block ------------------------
            if self._RE_ENDIF.match(line):
                if current is not None:
                    yield self._finalize(current)
                    current = None
                continue

            # ---- if / elseif / else: flush previous, start new --------------
            control_match = self._RE_CONTROL.match(line)
            if control_match:
                saw_control = True
                # Flush the block we were building
                if current is not None:
                    yield self._finalize(current)

                keyword = control_match.group(1).lower()
                current = self._start_block(keyword, line)
                continue

            # ---- action / set line: accumulate into active block ------------
            if current is not None:
                # Guard against stray keywords embedded mid-line
                if not any(lower.startswith(t) for t in self._TERMINATORS):
                    current["action_lines"].append(line)
            
            else:
                # No active block yet — stash for potential standalone emission
                standalone_actions.append(line)

        # ---- End-of-list: flush trailing block (handles missing endif) ------
        if current is not None:
            yield self._finalize(current)
        
        
        # ✅ Handle policies with NO control keywords at all
        if not saw_control and standalone_actions:
            yield {
                "condition": "",
                "dest_in": "",
                "other_match": "",
                "actions": ", ".join(standalone_actions),
            }
            
    
    def _start_block(self, keyword: str, line: str) -> dict:
        """Initialize a new condition block from a control line."""
        dest_in = ""
        other_match = ""

        if keyword in ("if", "elseif"):
            body_match = self._RE_CONDITION_BODY.match(line)
            body = body_match.group(1).strip() if body_match else ""

            dest_hit = self._RE_DEST_IN.search(body)
            if dest_hit:
                # dest_in = dest_hit.group(0).strip()   # full "destination in XYZ"
                
                # ✅ Use group(1) — just the group name, WITHOUT "destination in"
                dest_in = dest_hit.group(1).strip()
                
                # Remove the destination clause; whatever remains is "other"
                residual = self._RE_DEST_IN.sub("", body).strip(" ()&|,")
                other_match = residual
            
            else:
                other_match = body
        # `else` has no condition body — both match columns stay empty

        return {
            "condition": keyword,
            "dest_in": dest_in,
            "other_match": other_match,
            "action_lines": [],
        }
        
    
    @staticmethod
    def _finalize(block: dict) -> dict:
        """Convert an in-progress block into the output row shape."""
        return {
            "condition": block["condition"],
            "dest_in": block["dest_in"],
            "other_match": block["other_match"],
            "actions": ", ".join(block["action_lines"]),
        }
        
    
    @nw.narwhalify
    def value_normalizer_xr(self, df: IntoDataFrameT):
        df = df.with_columns(
            nw.col("value").cast(nw.String).str.replace("\r\n", "\n").str.replace("\r", "\n").str.strip_chars().str.split('\n').alias("condition_split_list")
        )
        native_df = df.to_native()
        if isinstance(native_df, pd.DataFrame):
            native_df["condition_split_list_stripped"] = native_df["condition_split_list"].apply(lambda x: [item.strip() for item in x])
        
        
        if isinstance(native_df, pl.DataFrame):
            native_df = native_df.with_columns(
                pl.col("condition_split_list").map_elements(lambda x: [item.strip() for item in x]).alias("condition_split_list_stripped")
            )
        
        df = nw.from_native(native_df)
        return df
    
    
    def section_writer(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, df: IntoDataFrameT, vendor_type: AnyStr="xr"):
        start_row = 1
        
        # worksheet = workbook_object["prefix"]
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row != 1 and max_col != 1:
            worksheet.delete_rows(1, max_row)
            
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if vendor_type == 'xr':
            neo_df = self.value_normalizer_xr(df)
                        
            if worksheet.max_row == 1 and worksheet.cell(1, 1).value is None:
                for element_id, element in enumerate(list(self._HEADERS), 1):
                    worksheet.cell(row=1, column=element_id, value=element)
            
            self.pattern_matcher_extracter_and_writer_xr(worksheet, neo_df)
            
    
            