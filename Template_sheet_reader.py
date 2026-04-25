import os
import pprint
import traceback
from queue import Queue
from turtle import width
import openpyxl
from excel_file_modifier import ExcelReader, ExcelSheetModifier
import pandas as pd
from collections import defaultdict
from threading import Event
from concurrent.futures import ThreadPoolExecutor

class Excel_Reader_and_Template_Maker:
    def __init__(self, file_path: str):
        try:
            if os.path.exists(str(file_path)):
                self.file_path = str(file_path)
                self.dataFrame = None
                self.dict = None
                self.excel_modifier_object = None
                self.sheets = None
                self.main_sections = None
                self.excel_reader_object = None
                self.workbook_loaded = None
                self.excel_modifier_dict = {}
                self.queue = Queue()
                
                if self.sheets is None:
                    if self.excel_reader_object is None:
                        self.excel_reader_object = ExcelReader(self.file_path)
                        
                
                self.sheets = self.excel_reader_object.get_sheets
                
                if self.sheets is None:
                    raise Exception("Sheets Not Found!")
                
            
            else:
                raise FileNotFoundError("File Not Found!")
        
        except FileNotFoundError:
            print("Error: File Not Found!")
        
        except Exception as e:
            print(f"Error: {traceback.format_exc()}\n{e}")
            
    def section_data_parser(self, sheet_name: str, section_name: str, start_row: int, end_row: int, max_sheet_column: int, main_start_column: int, ws: openpyxl.worksheet.worksheet.Worksheet):
        stop_event = Event()
        section_data = defaultdict(list)
        
        max_column_section = 1
        
        i = main_start_column + 1
        while i <= max_sheet_column:
            if ws.cell(row=start_row, column=i).value is not None:
                max_column_section = i
            
            else:
                break
            i += 1
        
        columns = [
            str(ws.cell(row=start_row, column=i).value).strip()
            for i in range(main_start_column+1, max_column_section+1)
        ]
        
        i = start_row + 1
        while i < end_row:
            if all(
                map(
                            lambda x: ws.cell(row=i, column=x).value is None,
                            range(main_start_column+1, len(columns)+1)
                        )
            ):
                i += 1
            else:
                j = 0
                while j < len(columns):
                    column_name = columns[j]
                    section_data[column_name].append(ws.cell(row=i, column=j+main_start_column+1).value)
                    j += 1
            i += 1
        
        dataframe = pd.DataFrame(section_data)
        result = (sheet_name, section_name, dataframe)
        
        self.queue.put(result)

        if not stop_event.is_set():
            stop_event.set()
    
    
    def excel_parser(self, sheet_name: str, excel_modifier_object: ExcelSheetModifier):
        stop_event = Event()
        try:
            __modifier_object = excel_modifier_object
            __sheet = __modifier_object.sheet
            main_start_row = __modifier_object.header_row
            main_start_column = __modifier_object.header_column
            
            max_row = __modifier_object.rows
            max_column = __modifier_object.columns
            _sections = []
            _section_start_row_dict = {}
            
            i = main_start_row
            
            # Appending the first Section in the Sections List
            _sections.append(str(__sheet.cell(row=i, column=main_start_column).value))
            _section_start_row_dict[str(__sheet.cell(row=i, column=main_start_column).value)] = i

            
            # Looping through the rows to find the next Sections
            while i <= max_row:
                # Finding the row in the excel sheet where all cells are empty and then finding the next section
                if all(
                    map(
                        lambda x: __sheet.cell(row=i, column=x).value is None,
                        range(main_start_column, max_column + 1)
                    )
                ):
                    while all(
                        map(
                            lambda x: __sheet.cell(row=i, column=x).value is None,
                            range(main_start_column, max_column + 1)
                        )
                    ):
                        i += 1
                    _sections.append(str(__sheet.cell(row=i, column=main_start_column).value))
                    _section_start_row_dict[str(__sheet.cell(row=i, column=main_start_column).value)] = i
                i += 1
                
                
            # i = 0
            # while i < len(_sections):
            #     selected_section = _sections[i]
            #     selected_section_start_row = _section_start_row_dict[selected_section]
            #     selected_section_end_row = _section_start_row_dict[_sections[i + 1]] if i + 1 < len(_sections) else max_row
                
            #     self.section_data_parser(selected_section, selected_section_start_row, selected_section_end_row, max_column, main_start_column, __sheet)
            #     i += 1
            with ThreadPoolExecutor(max_workers=4) as executor:
                    futures = [executor.submit(
                        self.section_data_parser,
                        sheet_name,
                        _sections[i],
                        _section_start_row_dict[_sections[i]],
                        _section_start_row_dict[_sections[i + 1]] if i + 1 < len(_sections) else max_row,
                        max_column,
                        main_start_column,
                        __sheet
                    )
                    for i in range(len(_sections))]
                    
                    for furture in futures:
                        furture.result()
        
        except Exception as e:
            print(f"Error: {traceback.format_exc()}\n{e}")
    
        if not stop_event.is_set():
            stop_event.set()
    
            
            
    def file_parser(self):
        try:
            self.workbook_loaded = self.excel_reader_object.get_openpyxl_workbook()
            if self.sheets:
                i = 0
                while i < len(self.sheets):
                    sheet_name = self.sheets[i]
                    # print(sheet_name)
                    excel_file_modifier = ExcelSheetModifier(self.workbook_loaded[sheet_name])       
                    self.excel_modifier_dict[sheet_name] = excel_file_modifier
                    i += 1
                    
                with ThreadPoolExecutor(max_workers=4) as executor:
                    futures = [
                        executor.submit(self.excel_parser, sheet_name, excel_file_modifier)
                        for sheet_name, excel_file_modifier in self.excel_modifier_dict.items()
                    ]
                    
                    for future in futures:
                        future.result()
            
            while not self.queue.empty():
                sheet_name, section_name, dataframe = self.queue.get()
                
                if self.dict is None:
                    self.dict = {}
                
                if sheet_name not in self.dict:
                    self.dict[sheet_name] = {}
                
                self.dict[sheet_name][section_name] = dataframe
                        
        except Exception as e:
            print(f"Error: {traceback.format_exc()}\n{e}")
        
    
    @property
    def get_dict(self):
        return self.dict if self.dict is not None else None
    
    def writer(self):
        try:
            string_ = f"{'\n\n'.join(f'{sheet_name}:\n\t{section_name}:\n{dataframe.to_markdown()}' for sheet_name, dict_ in self.dict.items() for section_name, dataframe in dict_.items())}"
            with open("test_text.txt", "w+") as f:
                f.write(string_)
                f.close()
                del f
        except Exception as e:
            print(f"Error: {traceback.format_exc()}\n{e}")
    
    def quit(self):
        if self.excel_modifier_list:
            self.excel_modifier_object.save( )
            del self.excel_modifier_object
            self.excel_modifier_object = None
        
        if self.excel_reader_object:
            self.excel_reader_object.quit()
            del self.excel_reader_object
            self.excel_reader_object = None