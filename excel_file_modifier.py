import os
import pandas as pd
import polars as pl
from typing import List, Optional
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from messages import Messagebox
from typing import Dict, AnyStr
import sections_writer as sw


class ExcelModifier:
    def __init__(
        self,
        workbook_path,
        sheet_name: str,
        dataframe: Optional[pd.DataFrame] = None,
        index_required: Optional[bool] = False,
        wrap_text: Optional[bool] = False,
    ):
        self.workbook = None
        self.workbook_to_be_saved = workbook_path

        if dataframe is not None:
            excel_writer = None
            if os.path.exists(workbook_path):
                excel_writer = pd.ExcelWriter(
                    workbook_path,
                    engine="openpyxl",
                    mode="a",
                    if_sheet_exists="replace",
                )

            else:
                excel_writer = pd.ExcelWriter(
                    workbook_path,
                    engine="openpyxl",
                    mode="w",
                    if_sheet_exists="replace",
                )

            if excel_writer:
                dataframe.to_excel(
                    excel_writer, sheet_name=sheet_name, index=index_required
                )
                excel_writer.save()
                excel_writer.close()
                del excel_writer

        self.worksheet = None
        self.columns = None
        self.rows = None
        self.header_row = None
        self.header_column = None

        self.side = Side(style="medium", color="000000")
        self.border = Border(
            left=self.side, right=self.side, top=self.side, bottom=self.side
        )
        self.fill = None
        self.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=wrap_text
        )
        self.header_font = Font(
            name="Ericsson Hilda", bold=True, size=14, color="FFFFFF"
        )
        self.normal_font = Font(name="Ericsson Hilda", size=11)

        if os.path.exists(workbook_path):
            self.workbook = load_workbook(workbook_path, read_only=False)
            self.worksheet = self.workbook[sheet_name]
            self.columns = self.worksheet.max_column
            self.rows = self.worksheet.max_row
            self.header_row, self.header_column = self.first_row_finder_for_header()

    def first_row_finder_for_header(self):
        i = 1
        j = 1
        breaker = False
        while i <= self.rows:
            while j <= self.columns:
                if self.worksheet.cell(row=i, column=j).value is not None:
                    header_row = i
                    header_column = j
                    breaker = True
                    break
                j += 1
            if breaker:
                break
            i += 1
        return header_row, header_column

    def column_width_adjuster(self):
        col_width = []
        default_font_size = 11

        i = self.header_row
        # print(f"{self.rows =}")

        total_columns = self.columns - self.header_column + 1

        while i <= self.rows:
            # print(f"{i = }")
            j = self.header_column
            while j <= self.columns:
                # print(f"{j = }")
                # print(f"{str(self.worksheet.cell(row=i, column=j).value) = }")

                cell_content_size = len(str(self.worksheet.cell(row=i, column=j).value).strip())
                cell = self.worksheet.cell(row=i, column=j)
                current_font_size = cell.font.size if cell.font and cell.font.size else default_font_size
                scaling_factor = current_font_size / default_font_size
                bold_multiplier = 1.45 if cell.font and cell.font.bold else 1.3
                required_size = (cell_content_size * scaling_factor * bold_multiplier)
                # print(f"{required_size =}")

                if len(col_width) < total_columns:

                    col_width.append(
                        required_size
                    )
                else:
                    list_index_to_be_updated = self.columns - (total_columns + j)
                    # print(f"{list_index_to_be_updated =}")
                    col_width[list_index_to_be_updated] = min(
                        max(col_width[list_index_to_be_updated], required_size), 50
                    )
                    # print(f"{col_width[list_index_to_be_updated] =}")

                j += 1
            i += 1

        i = 0
        while i < len(col_width):
            if col_width[i] < 50:
                col_width[i] += 3
            i += 1

        j = self.header_column
        while j <= self.columns:
            self.worksheet.column_dimensions[get_column_letter(j)].width = col_width[
                j - self.header_column
            ]
            j += 1

    def normal_styler(self):
        self.fill = PatternFill(
            start_color="3333FF", end_color="3333FF", fill_type="solid"
        )
        i = self.header_column
        while i <= self.columns:
            self.worksheet.cell(row=self.header_row, column=i).fill = self.fill
            self.worksheet.cell(
                row=self.header_row, column=i
            ).alignment = self.alignment
            self.worksheet.cell(row=self.header_row, column=i).font = self.header_font
            self.worksheet.cell(row=self.header_row, column=i).border = self.border
            i += 1

        i = self.header_row + 1
        while i <= self.rows:
            j = self.header_column
            while j <= self.columns:
                self.worksheet.cell(row=i, column=j).alignment = self.alignment
                self.worksheet.cell(row=i, column=j).border = self.border
                self.worksheet.cell(row=i, column=j).font = self.normal_font
                j += 1
            i += 1

        self.column_width_adjuster()
        self.save()

    def special_styler(self):
        pass

    def merger(self, range: List = None):
        pass
    
    @property
    def get_openpyxl_workbook(self)->Workbook|None:
        return self.workbook

    def save(self):
        if self.workbook:
            self.workbook.save(self.workbook_to_be_saved)
            self.workbook.close()
            del self.workbook
            
            

class ExcelReader:
    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self.workbook_load = None
        self.sheets = None
        
    @property
    def get_sheets(self) -> List[str]|None:
        if os.path.exists(self.workbook_path):
            self.workbook_load = load_workbook(self.workbook_path, read_only=False)
            self.sheets = self.workbook_load.sheetnames
            return self.sheets
        else:
            messagebox = Messagebox()
            messagebox.showerror("File Not Found", "File Not Found!")
            return None
    
    def get_openpyxl_workbook(self)->Workbook|None:
        return self.workbook_load

    def save(self):
        if self.workbook_load:
            self.workbook_load.save(self.workbook_path)
            self.quit()
            
    def quit(self):
        if self.workbook_load:
            self.workbook_load.close()
            del self.workbook_load
            

class ExcelSheetModifier:
    def __init__(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        self.worksheet = ws
        self.header_row = None
        self.header_column = None
        self.columns = None
        self.rows = None
        
        self.columns = self.worksheet.max_column
        self.rows = self.worksheet.max_row
        self.header_row, self.header_column = self.first_row_finder_for_header()
    

    @property
    def sheet(self)->openpyxl.worksheet.worksheet.Worksheet:
        return self.worksheet
    
    
    def first_row_finder_for_header(self):
        header_row = 1
        header_column = 1
        i = 1
        j = 1
        breaker = False
        while i <= self.rows:
            while j <= self.columns:
                if self.worksheet.cell(row=i, column=j).value is not None:
                    header_row = i
                    header_column = j
                    breaker = True
                    break
                j += 1
            if breaker:
                break
            i += 1
            
        return header_row, header_column

    def column_width_adjuster(self):
        col_width = []
        default_font_size = 11

        i = self.header_row
        # print(f"{self.rows =}")

        total_columns = self.columns - self.header_column + 1

        while i <= self.rows:
            # print(f"{i = }")
            j = self.header_column
            while j <= self.columns:
                # print(f"{j = }")
                # print(f"{str(self.worksheet.cell(row=i, column=j).value) = }")

                cell_content_size = len(str(self.worksheet.cell(row=i, column=j).value).strip())
                cell = self.worksheet.cell(row=i, column=j)
                current_font_size = cell.font.size if cell.font and cell.font.size else default_font_size
                scaling_factor = current_font_size / default_font_size
                bold_multiplier = 1.45 if cell.font and cell.font.bold else 1.3
                required_size = (cell_content_size * scaling_factor * bold_multiplier)
                # print(f"{required_size =}")

                if len(col_width) < total_columns:

                    col_width.append(
                        required_size
                    )
                else:
                    list_index_to_be_updated = self.columns - (total_columns + j)
                    # print(f"{list_index_to_be_updated =}")
                    col_width[list_index_to_be_updated] = min(
                        max(col_width[list_index_to_be_updated], required_size), 50
                    )
                    # print(f"{col_width[list_index_to_be_updated] =}")

                j += 1
            i += 1

        i = 0
        while i < len(col_width):
            if col_width[i] < 50:
                col_width[i] += 3
            i += 1

        j = self.header_column
        while j <= self.columns:
            self.worksheet.column_dimensions[get_column_letter(j)].width = col_width[
                j - self.header_column
            ]
            j += 1

    def normal_styler(self, wrap_text: bool = False):
        self.side = Side(style="medium", color="000000")
        self.border = Border(
            left=self.side, right=self.side, top=self.side, bottom=self.side
        )
        self.fill = None
        self.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=wrap_text
        )
        self.header_font = Font(
            name="Ericsson Hilda", bold=True, size=14, color="FFFFFF"
        )
        self.normal_font = Font(name="Ericsson Hilda", size=11)
        self.wrap_text = wrap_text
        self.fill = PatternFill(
            start_color="3333FF", end_color="3333FF", fill_type="solid"
        )
        i = self.header_column
        while i <= self.columns:
            self.worksheet.cell(row=self.header_row, column=i).fill = self.fill
            self.worksheet.cell(
                row=self.header_row, column=i
            ).alignment = self.alignment
            self.worksheet.cell(row=self.header_row, column=i).font = self.header_font
            self.worksheet.cell(row=self.header_row, column=i).border = self.border
            self.worksheet.cell(row=self.header_row, column=i).wrap_text = self.wrap_text
            i += 1

        i = self.header_row + 1
        while i <= self.rows:
            j = self.header_column
            while j <= self.columns:
                self.worksheet.cell(row=i, column=j).alignment = self.alignment
                self.worksheet.cell(row=i, column=j).border = self.border
                self.worksheet.cell(row=i, column=j).font = self.normal_font
                j += 1
            i += 1

        self.column_width_adjuster()
        # self.save()

class Excel_Writer_and_modifier:
    def __init__(self, host_details: str, workbook_path: str) -> None:
        self.host_details = host_details
        self.workbook_path = workbook_path
        
        if not os.path.exists(self.workbook_path):
            self.workbook = Workbook()
        else:
            self.workbook = load_workbook(self.workbook_path)
        self.header_row_index = 1
        self.max_rows = 0
        self.max_cols = 0
        
        self.side = Side(style="medium", color="000000")
        self.border = Border(
            left=self.side, right=self.side, top=self.side, bottom=self.side
        )
        self.fill = None
        self.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.header_font = Font(
            name="Ericsson Hilda", bold=True, size=14, color="FFFFFF"
        )
        self.normal_font = Font(name="Ericsson Hilda", size=11)
        self.wrap_text = True
        self.fill = PatternFill(
            start_color="3333FF", end_color="3333FF", fill_type="solid"
        )
        
        self.acceptible_sections = [
            "Prefix set Configuration", 
            "Route Policy Configuration"]
        
        # self.section_column_mapping = {
        #     "prefix" : {
        #         "prefix_set_name*":	"Prefix-list Name",
        #         "ip_subnet": "Prefix-IP",	
        #         "expression_ge": "",	
        #         "expression_le": "", 	
        #         "expression_eq": "",	
        #         "Operation": "Action"
        #     }
        # }
        # self.section_required_columns = {
        #     "prefix": ["Action","Sequence Action","Version","Prefix-list Name","Sequence Number","Permit/Deny","Prefix-IP","Length","Route Policy Mapping","Access-list Protocol","Access-list Source IP","Access-list source Wild Mask","Access-list Destination IP","Access-list destination Wild Mask"]
        # }
        self.acceptible_sections_class_object_creater = {
            "Prefix set Configuration" : sw.Prefix_section,
            "Route Policy Configuration" : sw.Policy_section
        }
        
        self.sheet_name_mapping = {
            "Route Policy Configuration": "policy",
            "Prefix set Configuration": "prefix"
        }
        
            
    def get_worksheet(self, sheet_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
        return self.workbook[sheet_name]
    
    def add_sheets(self, sheet_names: List[str]) -> None:
        for sheet_name in sheet_names:
            if sheet_name not in self.workbook.sheetnames:
                self.workbook.create_sheet(sheet_name)
        
        self.extra_sheet_remover(sheet_names)
        
    def styler(self, worksheet: openpyxl.worksheet.worksheet.Worksheet):
        max_cols = worksheet.max_column
        max_rows = worksheet.max_row
        col_width = []
        
        i = 1
        while i <= max_rows:
            j = 1
            while j <= max_cols:
                if len(col_width) < max_cols:
                    col_width.append(len(str(worksheet.cell(i, j).value)) + 3)
                else:
                    col_width[j-1] = max(col_width[j-1], len(str(worksheet.cell(i, j).value))) + 3
                j += 1
            i += 1
        
        for i in range(1, max_cols + 1):
            worksheet.column_dimensions[get_column_letter(i)].width = col_width[i-1]
            
        i = 1
        while i <= max_cols:
            worksheet.cell(1,i).fill = self.fill
            worksheet.cell(1,i).alignment = self.alignment
            worksheet.cell(1,i).font = self.header_font
            worksheet.cell(1,i).border = self.border
            i += 1
        
        i = 2
        while i <= max_rows:
            j = 1
            while j <= max_cols:
                worksheet.cell(i,j).border = self.border
                worksheet.cell(i,j).alignment = self.alignment
                worksheet.cell(i,j).font = self.normal_font
                j += 1
            i += 1
        
    
    def sheet_handler(self, dict_data: Dict[str, pl.DataFrame|pd.DataFrame], vendor_type: AnyStr|None) -> None:        
        sections = list(dict_data.keys())
        sheets_in_workbook = ["HostDetails"]
        i = 0
        while i < len(sections):
            selected_section = sections[i]
            selected_section_data = dict_data[selected_section]
            
            # start_row = 0
            # dest_section_name = next((x for x in self.acceptible_sections if x in str(selected_section).lower()), None)
            dest_section_name = self.sheet_name_mapping.get(selected_section, None)
            if dest_section_name:
                sheets_in_workbook.append(dest_section_name)
                if dest_section_name not in self.workbook.sheetnames:
                    self.workbook.create_sheet(dest_section_name)
                worksheet = self.workbook[dest_section_name]
                
                # max_row = worksheet.max_row
                # max_col = worksheet.max_column
                # columns = []
                
                # # Handle the section data
                # if isinstance(selected_section_data, pl.DataFrame):
                #     # Handle polars DataFrame
                #     columns = selected_section_data.columns
                
                # elif isinstance(selected_section_data, pd.DataFrame):
                #     # Handle pandas DataFrame
                #     columns = selected_section_data.columns.tolist()
                
                # if max_row in [0, 1, None]:
                #     start_row = 1
                # else:
                #     start_row = max_row + 2    
                # # worksheet.cell(row=start_row, column=1, value=selected_section)
                # # start_row += 1
                
                # # Add columns
                # required_columns = self.section_required_columns.get(dest_section_name, [])
                # # print(f"{required_columns = }\n")
                
                # col_dict = {col: idx for idx, col in enumerate(required_columns, 1)}
                # # print(f"{col_dict = }")
                # for column, col_idx in col_dict.items():
                #     worksheet.cell(row=start_row, column=col_idx, value=column)
                # start_row += 1
                # # print(f"{type(selected_section_data) = }")
                # # Add data
                # # if isinstance(selected_section_data, pl.DataFrame):
                # #     # Handle polars DataFrame
                # #     # for row_idx, row in enumerate(selected_section_data.iter_rows(), start=start_row):
                # #     #     for col_idx, value in enumerate(row, start=1):
                # #     #         worksheet.cell(row=row_idx, column=col_idx, value=value)
                    
                # #     selected_section_data.to_dict()
                
                # # elif isinstance(selected_section_data, pd.DataFrame):
                # #     # Handle pandas DataFrame
                # #     for row_idx, row in enumerate(selected_section_data.iterrows(), start=start_row):
                # #         for col_idx, value in enumerate(row[1], start=1):
                # #             worksheet.cell(row=row_idx, column=col_idx, value=value)
                
                # dict_ = selected_section_data.to_dict()
                # section_column_mapping = self.section_column_mapping[(dest_section_name).lower()]
                # # print(f"{section_column_mapping = }\n\n")
                # for key, value in dict_.items():
                #     col_id = col_dict.get(
                #         section_column_mapping.get(key, key), 1
                #     )
                #     for row_id in range(start_row, start_row + len(value) + 1):
                #         worksheet.cell(row=row_id, column=col_id, value=value[row_id - start_row - 1])
                
                self.acceptible_sections_class_object_creater[selected_section]().section_writer(worksheet, selected_section_data, vendor_type)
                self.styler(worksheet) 
            i += 1
        
        self.extra_sheet_remover(sheets_in_workbook)
        
    
    def add_data_from_dataframe_dict(self, data_dict: Dict[str, pl.DataFrame|pd.DataFrame], vendor_type: AnyStr='xr') -> None:
        # sheets = list(data_dict.keys())
        # self.add_sheets(sheets)
        # print(f"{data_dict.items() = }")0
        self.add_host_details_sheet()
        self.sheet_handler(data_dict, vendor_type)
    
    def add_host_details_sheet(self):
        if "HostDetails" not in self.workbook.sheetnames:
            self.workbook.create_sheet("HostDetails")
            host_details_sheet = self.workbook["HostDetails"]
            
            columns = ["Host Name","Host IP","CR ID","Vendor"]
            for i in range(len(columns)):
                host_details_sheet.cell(row=1, column=i+1, value=columns[i])
            host_details_sheet.cell(row=2, column=1, value=self.host_details)
            
    
    def quit(self)-> None:
        self.workbook.save(self.workbook_path)
        del self.workbook
    
    def extra_sheet_remover(self, sheet_names: List[str]) -> None:
        for sheet_name in self.workbook.sheetnames:
            if sheet_name not in sheet_names:
                self.workbook.remove(self.workbook[sheet_name])
                